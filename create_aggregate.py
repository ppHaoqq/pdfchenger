# メイン
import openpyxl
from pathlib import PurePath, Path
import string
from openpyxl.styles.borders import Border, Side
from statistics import median

excel_file = PurePath.joinpath(Path.cwd(), 'excel_datas')
name_file_ = PurePath.joinpath(excel_file, 'name_data')
file_list = list(name_file_.glob('*.xlsx'))

for excel_path in file_list:
    wb = openpyxl.load_workbook(excel_path)

    # 判定用にシートの一覧取得
    wss = wb.sheetnames
    # データがあるシート取得
    data_sheet = wb.worksheets[0]
    a = len(list(data_sheet.columns)[0]) - 1

    # 作業用シートがなければ作成
    if not '集計フォーム' in wss:
        wb.create_sheet(title='集計フォーム')
    # 作業用ワークシート取得
    ws = wb.worksheets[1]
    # 見出し書き込み
    ws['F1'] = '今期受注件数'
    ws['H1'] = '件'
    ws['F2'] = '今期受注金額'
    ws['H2'] = '円'

    ws['J1'] = '企業点'
    ws['L1'] = '点'
    ws['J2'] = '最高技術者評定点'
    ws['L2'] = '点'
    ws['J3'] = '評価点'
    ws['L3'] = '点'
    ws['J4'] = '平均入札率(過去{}件分)'.format(a)
    ws['L4'] = '%'
    ws['J5'] = '最低入札率(過去{}件分)'.format(a)
    ws['L5'] = '%'

    ws['A1'] = '技術者'
    ws['B1'] = '評点'
    ws['C1'] = '工期'
    ws['D1'] = '稼働状況'

    # 受注件数取得
    cell_o = list(data_sheet.columns)[16]
    # o列＝落差状況をリスト化、見出しを抜いたデータの個数で判断
    ws['G1'] = (len([o.value for o in cell_o if not o.value == None]) - 1)
    # 下記をリスト内包表記してる
    # o_list = []
    # for o in cell_o:
    #     if not o.value == None:
    #         o_list.append(o.value)
    # ws['B1'] = (len(o_list) - 1)

    # 受注金額取得
    # 落札決定、且つ、値が数値のものを抽出
    cell_l = list(data_sheet.columns)[12]
    # 金額合計用変数
    total_plice = 0
    # 判定のため2行展開
    for l_obj, o_obj in zip(cell_l, cell_o):
        l = l_obj.value.replace(',', '')
        o = o_obj.value
        # L行はカンマを削除後、数値判定（カンマ入りは全部Falseになる）
        # 条件を満たすものを合算
        if not o == None and l.isnumeric():
            total_plice += int(l)
    ws['G2'] = total_plice

    # 企業点取得
    subtotal1 = list(data_sheet.columns)[28]
    subtotal2 = list(data_sheet.columns)[30]
    date = list(data_sheet.columns)[8]
    # 日付の一覧をリスト化→最大値取得

    #     date_list = [d.value for d in date if not d.value=='入札日' and
    #                  subtotal1[d.row-1].value=='無効' or subtotal1[d.row-1].value=='小計1']
    date_list = []
    for d in date:
        if not d.value == '入札日':
            if not subtotal1[d.row - 1].value == '無効' or subtotal1[d.row - 1].value == '小計1':
                date_list.append(d.value)

    for aa_obj, ac_obj, date_obj in zip(subtotal1, subtotal2, date):
        cell_aa = aa_obj.value
        cell_ac = ac_obj.value
        d = date_obj.value
        # セルのデータが数字かどうか判定(STR.isnumeric())/日付が最大値と同じならOK
        if cell_aa.isnumeric() and cell_ac.isnumeric():
            e_score = int(cell_aa) + int(cell_ac)
            if d == max(date_list):
                ws['K1'] = e_score

    # 技術者評定点取得、仮技術者生成
    # 仮名用アルファベットリスト、社名用意
    al = list(string.ascii_uppercase)
    campany = data_sheet['H2'].value

    # 転記用データ取得
    t = list(data_sheet.columns)[21]

    # データ整形
    for t_obj in t:
        score = t_obj.value

        # 書き込み列取得, 空欄除去
        b_row = list(ws.columns)[1]
        b_row = [b for b in b_row if not b.value == None]
        # 列のうち、データが入っているセルの総数
        data_amount = len(b_row)
        # 総数-1で列の最後尾取得
        b_end = b_row[data_amount - 1].row

        # 数字のみフィルタリング
        if score.isnumeric():
            # すでに記載されている点数が同じ場合は除外＝同一人物とみなす
            # b_row 値を取得して再度リスト化
            if not int(score) in [r.value for r in b_row]:
                # 最後尾の一つ下のセルに入力
                ws.cell(row=b_end + 1, column=2, value=int(score))
                # 4列目を始点にアルファベットリストから取得、仮名作成
                engineer = campany + ':' + al[b_end - 1]
                # 最後尾の一つ下のセルに入力
                ws.cell(row=b_end + 1, column=1, value=engineer)

    # 最高技術者評定点取得
    # 既存の評点取得
    scores = list(ws.columns)[1]
    # 最大値判定用変数
    max_score = 0
    for s in scores:
        # 最大値がより出かければ更新
        if str(s.value).isnumeric():
            if int(s.value) > max_score:
                max_score = int(s.value)
    ws['K2'] = max_score

    # 評価点　計算
    #     ws['K3'] = (ws['K1'].value) + (ws['K2'].value)
    ws['K3'] = e_score + max_score

    # 入札率取得
    # 入札価格のみリスト化
    bid_list = list(data_sheet.columns)[12]

    # 列挿入
    if not data_sheet['N1'].value == '入札率':
        data_sheet.insert_cols(14)
        data_sheet['N1'] = '入札率'

    # 入札価格と行番号取得
    for p in bid_list:
        bid_val = p.value.replace(',', '')
        if bid_val.isnumeric():
            cell = 'D{}'.format(p.row)
            # 行番号をもとに該当する予定価格取得
            yotei = int(data_sheet[cell].value.replace('，', '').replace('円', ''))
            bid_rate = int(bid_val) / yotei
            write_cell = 'N{}'.format(p.row)
            # 該当する行番号に書き込み
            data_sheet[write_cell] = bid_rate
            # 集計フォームに集約、一覧取得後、中央値を表示
    # 　floot型のみリストに集計
    bid_rates = [br.value for br in list(data_sheet.columns)[13] if type(br.value) == float]
    # 中央値・最低値　算出後100倍して小数点2以下丸め
    med = round(median(bid_rates) * 100, 2)
    mini = round(min(bid_rates) * 100, 2)
    ws['K4'] = med
    ws['K5'] = mini

    # 罫線作成
    # 参考資料：https://qiita.com/github-nakasho/items/358e5602aeda81c58c81
    # 線のパラメーター作成
    # パラメの種類：https://chappy88.hatenablog.com/entry/2019/09/07/113405
    side = Side(style='thin', color='000000')
    # 線の位置を指定　Border(top=side, bottom=side, left=side, right=side)
    border = Border(side, side, side, side)

    # セル一つひとつに対して線作成を実行
    # シートから行を取り出す
    for row in ws:
        # 行からセルを取り出す
        for cell in row:
            # 値が入っているセルに対して式を実行
            if cell.value:
                cell.border = border

    # 幅調整
    for col in ws.columns:
        length = 0
        col_num = 0
        for cell in col:
            row = cell.row
            coo = cell.coordinate.rstrip(str(row))
            val = str(cell.value)
            if length < len(val):
                length = len(val)
            if not col_num == coo:
                col_num = coo
        col_num = str(col_num)
        ws.column_dimensions[col_num].width = (length + 6) * 1.2

    for col in data_sheet.columns:
        length = 0
        col_num = 0
        for cell in col:
            row = cell.row
            coo = cell.coordinate.rstrip(str(row))
            val = str(cell.value)
            if length < len(val):
                length = len(val)
            if not col_num == coo:
                col_num = coo
        col_num = str(col_num)
        data_sheet.column_dimensions[col_num].width = (length + 6) * 1.2

    # データ保存
    wb.save(excel_path)