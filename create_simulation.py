# データをもとにシュミレーションファイル作成

import openpyxl
from pathlib import PurePath, Path
import pandas as pd
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles.fonts import Font

excel_file = PurePath.joinpath(Path.cwd(), 'excel_datas')
name_file_ = PurePath.joinpath(excel_file, 'name_data')
file_list = list(name_file_.glob('*.xlsx'))

col = ['社名', '技術者', '技術者点', '企業点', '評価点', '加算点', '施工体制評価点', '基礎点', '持ち点',
       '予定価格＝積算価格', '低入率', '低入価格', '予想入札価格', '入札率', '評価値']
df = pd.DataFrame(index=[], columns=col)

k = input('工事名を入力してください：')
ip = input('積算価格を入力してください：')
if ',' in ip:
    ip = int(ip.replace(',', ''))
elif '，' in ip:
    ip = int(ip.replace('，', ''))
ip = int(ip)

fp = PurePath.joinpath(excel_file, '{}シュミレート.xlsx'.format(k))

for excel_path in file_list:
    wb = openpyxl.load_workbook(excel_path)
    # # 判定用にシートの一覧取得
    # wss = wb.sheetnames
    # # 作業用シートがなければ作成
    # if not '集計フォーム' in wss:
    #     wb.create_sheet(title='集計フォーム')
    # ワークシート取得
    ws = wb.worksheets[1]
    # data_sheet = wb.worksheets[0]

    name = excel_path.stem
    e_score = ws['K2'].value
    e_row = [s.row for s in list(ws.columns)[1] if s.value == e_score][0]
    engineer = ws['A{}'.format(e_row)].value
    c_score = ws['K1'].value
    eva_score = ws['K3'].value
    p_score = round(eva_score * 0.277, 1)
    con1 = 30
    con2 = 100
    score = (p_score + con1 + con2)
    integrated_price = ip
    con3 = 0.898
    low_price = int(integrated_price * con3)
    bid_rate = (ws['K4'].value) / 100
    bid_price = int(integrated_price * bid_rate)
    result = round(score / (bid_price / 100000000), 1)

    data_list = [name, engineer, e_score, c_score, eva_score, p_score, con1, con2, score, integrated_price,
                 con3, low_price, bid_price, bid_rate, result]

    sd = pd.Series(data_list, index=col)
    df = df.append(sd, ignore_index=True)

df.to_excel(fp, index=False)

# シュミレーションファイル整形

# excel_file = PurePath.joinpath(Path.cwd(), 'excel_datas')
# k = '横断道沖洲地区舗装'
# fp = PurePath.joinpath(excel_file, '{}シュミレート.xlsx'.format(k))

wb = openpyxl.load_workbook(fp)
ws = wb.active
rows = list(ws.rows)
cols = list(ws.columns)


# 罫線作成
side = Side(style='thin', color='000000')
border = Border(side, side, side, side)

# セル一つひとつに対して線作成を実行
for row in ws:
    # シートから行を取り出す
    #V 行からセルを取り出す
    for cell in row:
        # 値が入っているセルに対して式を実行
        if cell.value:
           cell.border = border

# ガイアートの文字色変更
font = Font(color='FF0000', italic=True)
n = 0
for col in cols[0]:
    if col.value == '（株）ガイアート':
        n += col.row
for cell in rows[n-1]:
    cell.font = font

# 評価値MAXに色付け

#　背景色参考資料：https://pg-chain.com/python-excel-color
# カラーコード一覧：http://www.netyasun.com/home/color.html
fill = PatternFill(patternType='solid', fgColor='FF6600')
font = Font(bold=True)
# 評価値一覧を取得、判定用列番号と合わせて辞書に格納
o_dict = {}
for o in cols[14]:
    row = o.row
    val = o.value
    if type(val)==float:
        o_dict[row] = val
# 辞書から最大値のキー取得https://note.nkmk.me/python-dict-value-max-min/
max_row = max(o_dict, key=o_dict.get)
# 列番号を使って該当列のセルに対して処理実行
for cell in rows[max_row-1]:
    cell.fill = fill
# 文字を太くする処理
    cell.font = font

# 金額列をカンマ.number_format = "#,##0"、少数を統一f.number_format = "#,##0.0"
for f, i, j, l, m in zip(cols[5], cols[8], cols[9], cols[11], cols[12]):
    if str(f.value).isnumeric() and str(i.value).isnumeric() and str(j.value).isnumeric() and str(l.value).isnumeric() and str(m.value).isnumeric():
        f.number_format = "#,##0.0"
        i.number_format = "#,##0.0"
        j.number_format = "#,##0"
        l.number_format = "#,##0"
        m.number_format = "#,##0"

# セル幅調整
for col in cols:
    length = 0
    col_num = 0
    for cell in col:
        row = cell.row
        coo = cell.coordinate.rstrip(str(row))
        val = str(cell.value)
        if length < len(val):
            length = len(val)
        if not col_num == coo:
            col_num = coo
    col_num = str(col_num)
    ws.column_dimensions[col_num].width = (length + 6) * 1.2

# 書式設定
# 評価値MAXに色付け
# from openpyxl.styles import PatternFill
# from openpyxl.styles.fonts import Font

# 　背景色参考資料：https://pg-chain.com/python-excel-color
# カラーコード一覧：http://www.netyasun.com/home/color.html
fill = PatternFill(patternType='solid', fgColor='FF6600')
font = Font(bold=True)
# 評価値一覧を取得、判定用列番号と合わせて辞書に格納
o_dict = {}
for o in cols[14]:
    row = o.row
    val = o.value
    if type(val) == float:
        o_dict[row] = val
# 辞書から最大値のキー取得https://note.nkmk.me/python-dict-value-max-min/
max_row = max(o_dict, key=o_dict.get)
# 列番号を使って該当列のセルに対して処理実行
for cell in rows[max_row - 1]:
    cell.fill = fill
    # 文字を太くする処理
    cell.font = font

# 式に変更
# 金額列をカンマ.number_format = "#,##0"、少数を統一f.number_format = "#,##0.0"
# 参考資料：https://pg-chain.com/python-excel-format
for e, f, i, j, l, m, o in zip(cols[4], cols[5], cols[8], cols[9], cols[11], cols[12], cols[14]):
    if (str(e.value).isnumeric() and type(f.value) == float or type(f.value) == int and
            type(i.value) == float or type(i.value) == int and str(l.value).isnumeric() and
            str(j.value).isnumeric() and type(m.value) == int and type(o.value) == float):
        ws[e.coordinate] = '=SUM(C{0}:D{0})'.format(e.row)
        ws[f.coordinate] = '=E{} * 0.277'.format(f.row)
        ws[i.coordinate] = '=SUM(F{0}:H{0})'.format(i.row)
        ws[m.coordinate] = '=J{0} * N{0}'.format(m.row)
        ws[o.coordinate] = '=(I{0} / (M{0}/100000000))'.format(o.row)

        ws[f.coordinate].number_format = "#,##0.0"
        ws[i.coordinate].number_format = "#,##0.0"
        ws[j.coordinate].number_format = "#,##0"
        ws[l.coordinate].number_format = "#,##0.0"
        ws[m.coordinate].number_format = "#,##0"
        ws[o.coordinate].number_format = "#,##0.0"

wb.save(fp)